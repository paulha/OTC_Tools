from jira_class import Jira, get_query
from navigate import *
import sys
from os.path import expanduser, pathsep, dirname, realpath
from openpyxl import load_workbook
from jira_class import Jira
from jira.exceptions import JIRAError
import re
from utility_funcs.search import search_for_file

class XLS:
    SUMMARY = 1
    DESCRIPTION = 6

    def __init__(self, filename, log=None):
        self.workbook = load_workbook(filename)
        self.log = log

    def headings(self, sheet='Sheet1'):
        row = self.workbook[sheet][1]
        return row

    def read(self, sheet='Sheet1'):
        active_heading = self.headings(sheet=sheet)
        active_sheet = self.workbook[sheet]
        result = []
        for row in range(2, active_sheet.max_row):
            this_entry = {'ROW': row}
            this_row = active_sheet[row]
            for col in range(0, this_row.__len__()):
                this_entry[active_heading[col].value] = this_row[col].value
            result.append(this_entry)
        return result


class Supersede:
    def __init__(self, jira, parser=None, scenario=None, config=None, queries=None, search=None, log=None):
        self.jira = jira
        self.parser = parser
        self.scenario = scenario
        self.config = config
        self.queries = queries
        self.search = search
        self.log = log
        self.update = True if 'update' in scenario and scenario['update'] else False
        self.areq = None
        self.preq = None
        self.row = 0

    def check_access(self, items):
        self.row = items['ROW']
        self.areq = self.jira.jira_client.issue(items['AREQ'])
        if self.areq is None:
            self.log.logger.error("Unable to access issue %s at row %s", self.areq.key, self.row)

        self.preq = self.jira.jira_client.issue(items['PREQ'])
        if self.preq is None:
            self.log.logger.error("Unable to access issue %s at row %s", self.preq.key, self.row)

        return True if self.areq is not None and self.preq is not None else False

    def copy_areq_values_to_preq(self):
        """Copy areq to preq, overlaying data if present"""

        update_occured = 0

        # Utility function for copying *_on fields (see below)
        def _define_update(update_list, field, entry):
            update_list[field] = [{'value': x.value} for x in getattr(entry.fields, field)] \
                if getattr(entry, field) is not None else []

        exists_on = self.jira.get_field_name('Exists On')
        verified_on = self.jira.get_field_name('Verified On')
        failed_on = self.jira.get_field_name('Failed On')
        blocked_on = self.jira.get_field_name('Blocked On')
        tested_on = self.jira.get_field_name('Tested On')
        validation_lead = self.jira.get_field_name('Validation Lead')
        classification = self.jira.get_field_name('Classification')

        val_lead = getattr(self.areq.fields, validation_lead)

        update_assignee_dict = {
            'assignee': {'name': self.areq.fields.assignee.name},
        }

        update_lead_dict = {
            validation_lead: {'name': val_lead.name if val_lead is not None else ""}
        }

        # -- Having created the issue, now other fields of the E-Feature can be updated:
        update_fields = {
            'priority': {'name': self.areq.fields.priority.name if self.areq is not None else 'P1-Stopper'},
            'labels': [x for x in getattr(self.areq.fields, 'labels')],
            'components': [{'id': x.id} for x in getattr(self.areq.fields, 'components')],
            classification: [{'id': x.id} for x in getattr(self.areq.fields, classification)]
        }
        _define_update(update_fields, verified_on, self.areq.fields)
        _define_update(update_fields, failed_on, self.areq.fields)
        _define_update(update_fields, blocked_on, self.areq.fields)
        _define_update(update_fields, tested_on, self.areq.fields)

        if 'exists_on' in self.scenario:
            target = self.scenario['exists_on']
        elif 'exists_only_on' in self.scenario:
            target = self.scenario['exists_only_on']
        else:
            target = None

        if target:
            exists_on_list = [{'value': target}]
            if 'exists_on' in self.scenario:
                exists_on_list.__add__([{'value': x.value} for x in getattr(self.areq.fields, exists_on)])

            update_fields[exists_on] = exists_on_list

        self.log.logger.debug("Updating E-Feature values from %s to %s" % (self.areq.key, self.preq.key))

        # -- Create the e-feature and update the stuff you can't set directly
        if self.update:
            self.preq.update(notify=False, fields=update_fields)
            try:
                self.preq.update(notify=False, fields=update_assignee_dict)
            except JIRAError as e:
                self.log.logger.error("Jira error %s", e)

            try:
                self.preq.update(notify=False, fields=update_lead_dict)
            except JIRAError as e:
                self.log.logger.error("Jira error %s", e)

            # -- Add a comment noting the creation of this feature.
            self.jira.jira_client.add_comment(self.preq,
                                              """This E-Feature was created by {command}.
    
                                              Parent Feature is %s. Source sibling is %s
                                              Source Platform: '{splatform}' Version '{sversion}'
    
                                              %s""".format_map(self.scenario)
                                              % (self.preq.key, self.areq.key,
                                                 self.scenario['comment'] if self.scenario['comment']
                                                                             is not None else ""))
            update_occured += 1

            self.log.logger.info("Updated PREQ %s for from %s: ", self.preq.key, self.areq.key)

        return update_occured > 0

    def create_areq_duplicate_of_preq_link(self):
        update_count = 0

        # -- Check to see if there's a link between this areq and preq
        has_duplicates_link = [link.outwardIssue
                               for link in self.areq.fields.issuelinks
                               if hasattr(link, "outwardIssue")
                                  and link.type.name == "Duplicates"
                                  and self.areq.fields.issuelinks[0].outwardIssue.key == self.preq.key
                               ]

        if not has_duplicates_link:
            if self.update:
                self.jira.create_issue_link("Duplicate", self.preq, self.areq,
                                            comment={"body": "Duplicate Feature link added from %s to %s"
                                                     % (self.preq.key, self.areq.key)})
                self.log.logger.info("Created 'Duplicates' link: %s --> %s", self.areq.key, self.preq.key)
                updated = True
            else:
                self.log.logger.warning("Link from %s --> %s is MISSING", self.areq.key, self.preq.key)

        if self.update:
            update_count += 1

        return update_count > 0

    def set_areq_to_Rejected(self):
        update_occured = 0
        if self.update:
            StateMachine.transition_to_state(self.jira, self.areq, E_Feature_Reject, self.log)
            update_occured += 1

        return update_occured > 0

    def supersede(self, items):
        updated = False
        if not self.check_access(items):
            self.log.logger.error("Abandoning row...")
            return False
        updated = updated or self.copy_areq_values_to_preq()
        updated = updated or self.create_areq_duplicate_of_preq_link()
        updated = updated or self.set_areq_to_Rejected()

        return updated




def areq_superceded_by_preq(parser, scenario, config, queries, search, log=None):
    """Copy values from areq to preq, Link areq to preq as 'Duplicates', then Reject areq, leave comment on areq and preq.

    Also:

    * Make sure areq and preq don't already have a 'Duplicates' link!
    * If areq status is already 'Reject', no need to re-set it.
    * If corresponding values are already set, no need to re-set them.
    * Should target values be overwritten if they are set (kinda think not...)
    * If areq status fails to be 'Reject', flag clearly!
    """

    update = scenario['update']
    log.logger.info("Update is %s", update)
    log.logger.info("Marking AREQ items as superseded by PREQ item.")
    log.logger.info("=================================================================")

    updates = 0

    XLS_FILE = realpath(dirname(realpath(sys.argv[0])) + '/../' + scenario['xls_input'])
    xls = XLS(XLS_FILE)
    items = xls.read()

    # -- Get and format it:
    jira = Jira(scenario['name'], search, log=log.logger)

    issue = jira.jira_client.issue("AREQ-23223")

    supersede = Supersede(jira, parser, scenario, config, queries, search, log=log)
    updates = 0
    for item in items:
        log.logger.info("Processing %s, %s and %s", item['ROW'], item['AREQ'], item['PREQ'])
        changed = supersede.supersede(item)

        updates = updates + 1 if changed else updates

        if scenario['createmax'] and updates >= scenario['createmax']:
            break

    log.logger.info("-----------------------------------------------------------------")
    log.logger.info("%s items were updated ", updates)
    log.logger.info("")

    return


